import csv
import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import aliased
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import get_db
from app.dependencies import require_admin, require_superadmin, is_scoped, program_scope_condition
from app.models.auth import User
from app.models.response import Response, ResponseItem
from app.models.respondent import Respondent
from app.models.instrument import CippDimension, CippSubDimension, InstrumentItem
from app.models.academic import Course, Program, Faculty

router = APIRouter(prefix="/admin/export", tags=["admin-export"])

HEADER_FILL = PatternFill("solid", fgColor="1A365D")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _scope_responses(q, current_user: User):
    if is_scoped(current_user):
        q = q.join(Course, Response.course_id == Course.id).join(
            Program, Course.program_id == Program.id
        ).where(program_scope_condition(current_user))
    return q


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_response(rows: list[list], filename: str) -> StreamingResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 18


# ══════════════ EXPORT RESPONS (UC-18) ════════════════════════════════════════

@router.get("/responses")
async def export_responses(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    course_id: Optional[int] = None,
    role: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_superadmin),
):
    """Export semua respons lengkap beserta profil responden.

    UC-18d: data mentah individual hanya untuk superadmin.
    """
    # Fakultas diturunkan dari MK → prodi → fakultas (berlaku untuk mahasiswa & dosen)
    Lecturer = aliased(User)
    q = (
        select(Response, Respondent, Course, Faculty, Lecturer)
        .join(Course, Response.course_id == Course.id)
        .join(Program, Course.program_id == Program.id)
        .outerjoin(Faculty, Program.faculty_id == Faculty.id)
        .outerjoin(Lecturer, Response.evaluated_lecturer_id == Lecturer.id)
        .where(Response.is_complete == True)
    )
    if is_scoped(current_user):
        q = q.where(program_scope_condition(current_user))
    q = q.outerjoin(Respondent, Response.respondent_id == Respondent.id)
    if course_id:
        q = q.where(Response.course_id == course_id)
    if role:
        q = q.where(Response.role == role)
    q = q.order_by(Response.submitted_at.desc())
    rows_db = (await db.execute(q)).all()

    headers = [
        "ID Respons", "Role", "Mata Kuliah", "Fakultas", "Dosen Dinilai", "Bahasa",
        "Nama Responden", "Jenis Kelamin", "Usia", "Semester",
        "Lama Belajar Mandarin", "Level HSK", "Pernah ke China",
        "Tanggal Submit",
    ]
    data_rows = []
    for resp, respondent, course, faculty, lecturer in rows_db:
        faculty_name = faculty.nama if faculty else (respondent.faculty if respondent else "")
        data_rows.append([
            str(resp.id),
            resp.role,
            f"{course.kode_mk} — {course.nama_id}",
            faculty_name or "",
            lecturer.full_name if lecturer else "",
            resp.bahasa,
            respondent.full_name if respondent else resp.role,
            respondent.gender if respondent else "",
            respondent.age if respondent else "",
            respondent.current_semester if respondent else "",
            respondent.mandarin_study_duration if respondent else "",
            respondent.hsk_level_mahasiswa if respondent else "",
            respondent.china_stay_duration if respondent else "",
            resp.submitted_at.strftime("%Y-%m-%d %H:%M") if resp.submitted_at else "",
        ])

    if format == "csv":
        return _csv_response([headers] + data_rows, "sinoku_responses.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Respons"
    _style_header(ws, headers)
    for row in data_rows:
        ws.append(row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    return _xlsx_response(wb, "sinoku_responses.xlsx")


# ══════════════ EXPORT SKOR CIPP (UC-18) ══════════════════════════════════════

@router.get("/scores")
async def export_scores(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    course_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Export skor rata-rata CIPP per item, dipisah dosen dan mahasiswa."""
    q_resp = select(Response).where(Response.is_complete == True)
    q_resp = _scope_responses(q_resp, current_user)
    if course_id:
        q_resp = q_resp.where(Response.course_id == course_id)
    responses = (await db.execute(q_resp)).scalars().all()

    ids_dosen = [r.id for r in responses if r.role == "dosen"]
    ids_mhs = [r.id for r in responses if r.role == "mahasiswa"]

    # Ambil semua item dengan dimensi dan subdimensi
    items_rows = (await db.execute(
        select(InstrumentItem, CippSubDimension, CippDimension)
        .join(CippSubDimension, InstrumentItem.sub_dimension_id == CippSubDimension.id)
        .join(CippDimension, CippSubDimension.dimension_id == CippDimension.id)
        .where(InstrumentItem.is_active == True)
        .order_by(CippDimension.urutan, CippSubDimension.urutan, InstrumentItem.nomor_urut)
    )).all()

    # Batch query skor
    all_resp_ids = ids_dosen + ids_mhs
    if all_resp_ids:
        score_rows = (await db.execute(
            select(ResponseItem.item_id, ResponseItem.response_id, ResponseItem.skor)
            .where(ResponseItem.response_id.in_(all_resp_ids))
        )).all()
    else:
        score_rows = []

    # Index skor: {item_id: {response_id: skor}}
    score_index: dict = {}
    for item_id, resp_id, skor in score_rows:
        score_index.setdefault(item_id, {})[resp_id] = skor

    def mean_scores(item_id, resp_ids):
        scores = [score_index.get(item_id, {}).get(rid) for rid in resp_ids]
        scores = [s for s in scores if s is not None]
        return round(sum(scores) / len(scores), 2) if scores else ""

    headers = [
        "Dimensi", "Sub-Dimensi", "Kode Item", "No. Urut",
        "Teks Item (Dosen)", "Teks Item (Mahasiswa)",
        "Rata-rata Dosen", "Rata-rata Mahasiswa", "N Dosen", "N Mahasiswa",
    ]
    data_rows = []
    for item, subdim, dim in items_rows:
        n_d = len([r for r in ids_dosen if score_index.get(item.id, {}).get(r) is not None])
        n_m = len([r for r in ids_mhs if score_index.get(item.id, {}).get(r) is not None])
        data_rows.append([
            dim.kode,
            subdim.kode,
            item.kode,
            item.nomor_urut,
            item.text_id_dosen,
            item.text_id_mahasiswa,
            mean_scores(item.id, ids_dosen),
            mean_scores(item.id, ids_mhs),
            n_d,
            n_m,
        ])

    if format == "csv":
        return _csv_response([headers] + data_rows, "sinoku_scores.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Skor CIPP"
    _style_header(ws, headers)
    for row in data_rows:
        ws.append(row)
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 50
    for col_letter in ["A", "B", "C", "D", "G", "H", "I", "J"]:
        ws.column_dimensions[col_letter].width = 16
    return _xlsx_response(wb, "sinoku_scores.xlsx")


# ══════════════ EXPORT PROFIL RESPONDEN (UC-18) ═══════════════════════════════

@router.get("/respondents")
async def export_respondents(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    course_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_superadmin),
):
    """Export profil lengkap responden mahasiswa.

    UC-18d: data mentah individual hanya untuk superadmin.
    """
    Lecturer = aliased(User)
    q = (
        select(Respondent, Response, Course, Faculty, Lecturer)
        .join(Response, Response.respondent_id == Respondent.id)
        .join(Course, Response.course_id == Course.id)
        .join(Program, Course.program_id == Program.id)
        .outerjoin(Faculty, Program.faculty_id == Faculty.id)
        .outerjoin(Lecturer, Response.evaluated_lecturer_id == Lecturer.id)
        .where(Response.is_complete == True, Response.role == "mahasiswa")
    )
    if is_scoped(current_user):
        q = q.where(program_scope_condition(current_user))
    if course_id:
        q = q.where(Response.course_id == course_id)
    q = q.order_by(Response.submitted_at.desc())
    rows_db = (await db.execute(q)).all()

    headers = [
        "ID Respons", "Nama", "Jenis Kelamin", "Usia", "Semester",
        "Fakultas", "Dosen Dinilai", "Lama Belajar Mandarin", "Level HSK",
        "Pernah ke China", "Memiliki Teman Tionghoa",
        "Status MK", "Frekuensi Interaksi Budaya",
        "Mata Kuliah", "Tanggal Submit",
    ]
    data_rows = []
    for respondent, resp, course, faculty, lecturer in rows_db:
        faculty_name = faculty.nama if faculty else (respondent.faculty or "")
        data_rows.append([
            str(resp.id),
            respondent.full_name or "(anonim)",
            respondent.gender or "",
            respondent.age or "",
            respondent.current_semester or "",
            faculty_name,
            lecturer.full_name if lecturer else "",
            respondent.mandarin_study_duration or "",
            respondent.hsk_level_mahasiswa or "",
            respondent.china_stay_duration or "",
            respondent.chinese_friends or "",
            respondent.course_status_taken or "",
            respondent.cultural_interaction_freq or "",
            f"{course.kode_mk} — {course.nama_id}",
            resp.submitted_at.strftime("%Y-%m-%d %H:%M") if resp.submitted_at else "",
        ])

    if format == "csv":
        return _csv_response([headers] + data_rows, "sinoku_respondents.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Profil Responden"
    _style_header(ws, headers)
    for row in data_rows:
        ws.append(row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    return _xlsx_response(wb, "sinoku_respondents.xlsx")
